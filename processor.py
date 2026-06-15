# ============================================================
# CELL 2 — MAIN PROCESSING CODE
# ============================================================

import pandas as pd
import os
import shutil
import re
from google.colab import files
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, Border, PatternFill
from tqdm.auto import tqdm

# ============================================================
# CHECK CONFIG
# ============================================================

if "CONFIG_READY" not in globals() or CONFIG_READY is not True:
    raise RuntimeError("Run CELL 1 first, choose settings, then click 🍇 SAVE CONFIG.")

if "PRINT_SAVED_LINES" not in globals():
    PRINT_SAVED_LINES = False

if "PRINT_FULL_SUMMARY_ON_SCREEN" not in globals():
    PRINT_FULL_SUMMARY_ON_SCREEN = False

# ============================================================
# FIXED FINAL KAUM BUCKETS
# ============================================================
# Final buckets only:
# MELAYU / CINA / INDIA / LAIN-LAIN
#
# ORANG ASLI = LAIN-LAIN
# BUMI = LAIN-LAIN
# anything else = LAIN-LAIN
# ============================================================

ORIGINAL_KAUM_BUCKETS = ["MELAYU", "CINA", "INDIA", "LAIN-LAIN"]

# ============================================================
# CLEAN OUTPUT FOLDER
# ============================================================

if os.path.exists(OUTPUT_ROOT):
    shutil.rmtree(OUTPUT_ROOT)

os.makedirs(OUTPUT_ROOT, exist_ok=True)

# ============================================================
# UPLOAD FILES
# ============================================================

uploaded = files.upload()

if not uploaded:
    raise ValueError("No files uploaded.")

# ============================================================
# COLUMN ALIASES
# ============================================================

COLUMN_ALIASES = {
    "name": [
        "name", "nama", "nama penuh"
    ],

    "nokp": [
        "nokp", "no kp", "no.kp", "no kad pengenalan",
        "ic", "no ic", "nric", "no k/p", "no kp lama"
    ],

    "jantina": [
        "jantina", "gender", "sex"
    ],

    "umur": [
        "umur", "age"
    ],

    "kaum_spr": [
        "kaum spr", "kaum_spr"
    ],

    "kategori_kaum": [
        "kategori kaum", "kategori_kaum", "kaum", "bangsa", "race"
    ],

    "number": [
        "number", "phone", "phone 1", "no tel", "no tel 1",
        "no telefon", "nombor", "telefon", "mobile phone",
        "mobile", "mobile no", "mobile number"
    ],

    "number2": [
        "phone 2", "no tel 2", "no telefon 2", "telefon 2"
    ],

    "kod_lokaliti": [
        "kod lokaliti", "kod_lokaliti", "locality code", "locality_code"
    ],

    "nama_lokaliti": [
        "nama lokaliti", "nama_lokaliti", "locality"
    ],

    "kod_dm": [
        "kod dm", "kod_dm", "dm code", "dm_code"
    ],

    "nama_dm": [
        "nama dm", "nama_dm", "dm"
    ],

    "kod_dun": [
        "kod dun", "kod_dun", "dun code", "dun_code"
    ],

    "nama_dun": [
        "nama dun", "nama_dun", "dun"
    ],

    "kod_parlimen": [
        "kod parlimen", "kod_parlimen", "parliament code", "parliament_code"
    ],

    "nama_parlimen": [
        "nama parlimen", "nama_parlimen", "parlimen", "parliament"
    ],

    "kod_negeri": [
        "kod negeri", "kod_negeri", "state code", "state_code"
    ],

    "nama_negeri": [
        "nama negeri", "nama_negeri", "negeri", "state"
    ],

    "sikap": [
        "sikap"
    ],

    "party": [
        "party", "parti"
    ],
}

# ============================================================
# BASIC FUNCTIONS
# ============================================================

def normalize_key(text):
    text = str(text)
    text = text.replace("\xa0", " ")
    text = text.replace("\n", " ")
    text = text.replace("\r", " ")
    text = text.replace("_", " ")
    text = text.strip().upper()
    text = re.sub(r"\s+", " ", text)
    return text


ALIAS_LOOKUP = {}

for canonical, aliases in COLUMN_ALIASES.items():
    for alias in aliases:
        ALIAS_LOOKUP[normalize_key(alias)] = canonical


def sanitize(text):
    text = str(text).strip().upper()
    text = re.sub(r"\s+", " ", text)
    return re.sub(r'[\\/*?:"<>|]', "_", text)


def make_unique_columns(cols):
    seen = {}
    final_cols = []

    for col in cols:
        norm = normalize_key(col)

        if norm in ALIAS_LOOKUP:
            new_col = ALIAS_LOOKUP[norm]
        else:
            new_col = str(col).strip()
            new_col = new_col.replace("\xa0", " ")
            new_col = re.sub(r"\s+", " ", new_col)

            if new_col == "" or new_col.upper().startswith("UNNAMED"):
                new_col = "UNNAMED"

        if new_col not in seen:
            seen[new_col] = 0
            final_cols.append(new_col)
        else:
            seen[new_col] += 1
            final_cols.append(f"{new_col}__{seen[new_col]}")

    return final_cols


def detect_header_row(raw_df):
    best_row = None
    best_score = 0
    max_scan = min(100, len(raw_df))

    for i in range(max_scan):
        row_values = [normalize_key(x) for x in raw_df.iloc[i].tolist()]
        score = sum(1 for x in row_values if x in ALIAS_LOOKUP)

        if score > best_score:
            best_score = score
            best_row = i

    if best_score >= 3:
        return best_row

    return None


def read_excel_smart(file_name):
    all_sheets = []

    xls = pd.ExcelFile(file_name)
    sheet_names = xls.sheet_names if READ_ALL_SHEETS else [xls.sheet_names[0]]

    print(f"\nReading workbook: {file_name}")
    print(f"Sheets used: {sheet_names}")

    for sheet_name in sheet_names:
        raw = pd.read_excel(
            file_name,
            sheet_name=sheet_name,
            header=None,
            dtype=str
        ).fillna("")

        header_row = detect_header_row(raw)

        if header_row is None:
            print(f"  Skipped sheet: {sheet_name} — header not detected")
            continue

        cols = raw.iloc[header_row].tolist()
        cols = make_unique_columns(cols)

        data = raw.iloc[header_row + 1:].copy()
        data.columns = cols
        data = data.fillna("")

        data = data[
            data.astype(str).apply(
                lambda row: "".join(row.values).strip() != "",
                axis=1
            )
        ].copy()

        if data.empty:
            print(f"  Skipped sheet: {sheet_name} — no rows")
            continue

        data["source_file"] = file_name
        data["source_sheet"] = sheet_name

        all_sheets.append(data)

        print(f"  Loaded sheet: {sheet_name} — {len(data):,} rows")

    if not all_sheets:
        raise ValueError(f"No usable sheet found in {file_name}")

    return pd.concat(all_sheets, ignore_index=True)


def clean_phone(value):
    phone = re.sub(r"\D", "", str(value).strip())

    if phone.startswith("60") and len(phone) in [11, 12]:
        phone = "0" + phone[2:]

    if phone == "":
        return None

    if "0000" in phone:
        return None

    if not phone.startswith("01"):
        return None

    if not phone.isdigit():
        return None

    if not (10 <= len(phone) <= 11):
        return None

    return phone


def choose_one_phone(row):
    for col in PHONE_PRIORITY_COLUMNS:
        if col in row:
            phone = clean_phone(row.get(col, ""))

            if phone is not None:
                return phone

    return None


def normalize_filter_value(x):
    x = str(x).strip().upper()
    x = x.replace("_", " ")
    x = re.sub(r"\s+", " ", x)

    if x in ["LAIN LAIN", "LAIN-LAIN", "LAINLAIN"]:
        return "LAIN-LAIN"

    if x == "ORANG ASLI":
        return "LAIN-LAIN"

    if "BUMI" in x:
        return "LAIN-LAIN"

    if x not in ["MELAYU", "CINA", "INDIA", "LAIN-LAIN"]:
        return "LAIN-LAIN"

    return x


def clean_kaum(value):
    k = str(value).strip().upper()
    k = k.replace("_", " ")
    k = re.sub(r"\s+", " ", k)

    if k == "MELAYU":
        return "MELAYU"

    elif k == "CINA":
        return "CINA"

    elif k == "INDIA":
        return "INDIA"

    else:
        return "LAIN-LAIN"


def normalize_sikap(value):
    s = str(value).strip().upper()

    s = s.replace("\xa0", " ")
    s = s.replace("_", " ")
    s = s.replace("–", "-")
    s = s.replace("—", "-")
    s = re.sub(r"\s+", " ", s)

    compact = re.sub(r"[^A-Z]", "", s)

    if "KELABU" in compact:
        return "KELABU"

    if "HITAM" in compact:
        return "HITAM"

    if "PUTIH" in compact:
        return "PUTIH"

    return s


def get_code(kategori_kaum, gender):
    k = clean_kaum(kategori_kaum)
    g = str(gender).strip().upper()

    if k == "MELAYU":
        return "ML" if g == "L" else "MP"

    elif k == "CINA":
        return "CL" if g == "L" else "CP"

    elif k == "INDIA":
        return "IL" if g == "L" else "IP"

    else:
        return "LLL" if g == "L" else "LLP"


def format_code_7digit(raw_code):
    raw = str(raw_code).strip()

    if "/" in raw:
        parts = raw.split("/")

        if len(parts) >= 3:
            parlimen = re.sub(r"\D", "", parts[0]).zfill(3)
            dun = re.sub(r"\D", "", parts[1]).zfill(2)
            last_raw = re.sub(r"\D", "", parts[2])
            last = str(int(last_raw)) if last_raw else "0"

            return f".{parlimen}.{dun}.{last}"

    digits = re.sub(r"\D", "", raw)

    if digits == "":
        return ""

    digits = digits.zfill(7)

    parlimen = digits[:3]
    dun = digits[3:5]
    last_raw = digits[5:7]
    last = str(int(last_raw)) if last_raw else "0"

    return f".{parlimen}.{dun}.{last}"


def format_first_name(row):
    return format_code_7digit(row.get("kod_dm", ""))


def assign_age_range(age):
    try:
        age = int(float(str(age).strip()))
    except:
        return None

    for txt_label, file_label, min_age, max_age in AGE_RANGES:
        if max_age is None:
            if age >= min_age:
                return txt_label
        else:
            if min_age <= age <= max_age:
                return txt_label

    return None


def normalize_parlimen_code(code):
    raw = str(code).strip().upper()
    num = re.sub(r"\D", "", raw)

    if num == "":
        return "P000"

    return f"P{int(num):03d}"


def extract_dun_num(kod_dun):
    raw = str(kod_dun).strip().upper()
    num = re.sub(r"\D", "", raw)

    if num == "":
        return "00"

    if len(num) >= 2:
        return str(int(num[-2:]))

    return str(int(num))


def extract_dm_num(kod_dm):
    raw = str(kod_dm).strip()

    if "/" in raw:
        parts = raw.split("/")

        if len(parts) >= 3:
            dm_raw = re.sub(r"\D", "", parts[2])
            return str(int(dm_raw)) if dm_raw else "0"

    digits = re.sub(r"\D", "", raw)

    if digits == "":
        return "0"

    digits = digits.zfill(7)
    dm_raw = digits[-2:]

    return str(int(dm_raw)) if dm_raw else "0"


def code_sort_num(text):
    num = re.sub(r"\D", "", str(text))
    return int(num) if num else 0


def format_count(n):
    return f"{int(n):,}"


def save_xlsx_no_style(df, path):
    wb = Workbook()
    ws = wb.active

    for row in dataframe_to_rows(df, index=False, header=True):
        ws.append(row)

    for row in ws.iter_rows():
        for cell in row:
            cell.font = Font(bold=False)
            cell.border = Border()
            cell.fill = PatternFill(fill_type=None)
            cell.alignment = Alignment(horizontal="left", vertical="top")

    wb.save(path)


def build_wa_df(group):
    return pd.DataFrame({
        "Title": group["number"].astype(str),
        "First Name": group["First Name"].astype(str),
        "Middle Name": "",
        "Last Name": group["Last Name"].astype(str),
        "Mobile Phone": group["number"].astype(str),
    })


def get_required_columns():
    req = set()

    req.add("number")
    req.add("kod_dm")
    req.add("kategori_kaum")
    req.add("jantina")

    for level in GROUP_LEVELS:
        if level == "PARLIMEN":
            req.update(["kod_parlimen", "nama_parlimen"])

        elif level == "DUN":
            req.update(["kod_dun", "nama_dun"])

        elif level == "DM":
            req.update(["kod_dm", "nama_dm"])

        elif level == "KAUM":
            req.add("kategori_kaum")

        elif level == "AGE":
            req.add("umur")

    if SIKAP_FILTER:
        req.add("sikap")

    if PARTY_FILTER:
        req.add("party")

    if AGE_FILTER is not None:
        req.add("umur")

    return sorted(req)


def validate_required_columns(df):
    req = get_required_columns()
    missing = []

    for col in req:
        if col == "number":
            if "number" not in df.columns and "number2" not in df.columns:
                missing.append("number / NO TEL 1")
        else:
            if col not in df.columns:
                missing.append(col)

    if missing:
        print("\nAvailable columns after cleanup:")
        print(list(df.columns))
        raise ValueError(f"Missing columns after header cleanup: {missing}")


def build_labels(df):
    if "kod_parlimen" in df.columns:
        df["__PARLIMEN_LABEL"] = df.apply(
            lambda x: f"{normalize_parlimen_code(x.get('kod_parlimen', ''))} {sanitize(x.get('nama_parlimen', ''))}",
            axis=1
        )
        df["__PARLIMEN_SORT"] = df["kod_parlimen"].apply(code_sort_num)

    if "kod_dun" in df.columns:
        df["__DUN_LABEL"] = df.apply(
            lambda x: f"N.{extract_dun_num(x.get('kod_dun', ''))} {sanitize(x.get('nama_dun', ''))}",
            axis=1
        )
        df["__DUN_SORT"] = df["kod_dun"].apply(code_sort_num)

    if "kod_dm" in df.columns:
        df["__DM_LABEL"] = df.apply(
            lambda x: f"DM{extract_dm_num(x.get('kod_dm', ''))} {sanitize(x.get('nama_dm', ''))}",
            axis=1
        )
        df["__DM_SORT"] = df["kod_dm"].apply(code_sort_num)

    if "kaum_clean" in df.columns:
        df["__KAUM_LABEL"] = df["kaum_clean"]

        kaum_order = {
            "MELAYU": 1,
            "CINA": 2,
            "INDIA": 3,
            "LAIN-LAIN": 4,
        }

        df["__KAUM_SORT"] = df["kaum_clean"].map(kaum_order).fillna(99).astype(int)

    if "umur" in df.columns:
        df["__AGE_LABEL"] = df["umur"].apply(assign_age_range)

        age_order = {
            txt_label: i
            for i, (txt_label, file_label, min_age, max_age) in enumerate(AGE_RANGES)
        }

        df["__AGE_SORT"] = df["__AGE_LABEL"].map(age_order).fillna(999).astype(int)

    return df


def get_group_columns():
    label_cols = []
    sort_cols = []

    for level in GROUP_LEVELS:
        if level == "PARLIMEN":
            label_cols.append("__PARLIMEN_LABEL")
            sort_cols.append("__PARLIMEN_SORT")

        elif level == "DUN":
            label_cols.append("__DUN_LABEL")
            sort_cols.append("__DUN_SORT")

        elif level == "DM":
            label_cols.append("__DM_LABEL")
            sort_cols.append("__DM_SORT")

        elif level == "KAUM":
            label_cols.append("__KAUM_LABEL")
            sort_cols.append("__KAUM_SORT")

        elif level == "AGE":
            label_cols.append("__AGE_LABEL")
            sort_cols.append("__AGE_SORT")

    return label_cols, sort_cols


def get_file_label_from_age(txt_label):
    for t, file_label, min_age, max_age in AGE_RANGES:
        if t == txt_label:
            return file_label

    return txt_label


def safe_file_label(label):
    label = str(label)

    if label in [x[0] for x in AGE_RANGES]:
        label = get_file_label_from_age(label)

    return sanitize(label)

# ============================================================
# READ ALL FILES
# ============================================================

all_data = []

for file_name in tqdm(uploaded.keys(), desc="Reading files"):
    df_file = read_excel_smart(file_name)
    all_data.append(df_file)

if not all_data:
    raise ValueError("No uploaded data found.")

df = pd.concat(all_data, ignore_index=True)

original_rows = len(df)

if "number2" not in df.columns:
    df["number2"] = ""

validate_required_columns(df)

print(f"\nOriginal rows             : {format_count(original_rows)}")

# ============================================================
# PHONE CLEANING
# ============================================================

df["number"] = df.apply(choose_one_phone, axis=1)

before_phone = len(df)
df = df[df["number"].notna()].copy()
after_phone = len(df)

print(f"After valid number filter : {format_count(after_phone)} kept ({format_count(before_phone - after_phone)} removed)")

if DEDUP_BY_NOKP and "nokp" in df.columns:
    df["nokp_clean"] = df["nokp"].astype(str).str.replace(r"\D", "", regex=True).str.strip()

    before_nokp = len(df)

    df_valid_nokp = df[df["nokp_clean"] != ""].drop_duplicates(
        subset=["nokp_clean"],
        keep="first"
    )

    df_blank_nokp = df[df["nokp_clean"] == ""]

    df = pd.concat([df_valid_nokp, df_blank_nokp], ignore_index=True)

    after_nokp = len(df)

    print(f"After NO KP duplicate cut : {format_count(after_nokp)} kept ({format_count(before_nokp - after_nokp)} removed)")

else:
    print("After NO KP duplicate cut : skipped, no NO KP column found")

if DEDUP_BY_PHONE:
    before_dup = len(df)
    df = df.drop_duplicates(subset=["number"], keep="first").copy()
    after_dup = len(df)

    print(f"After number duplicate cut: {format_count(after_dup)} kept ({format_count(before_dup - after_dup)} removed)")

if df.empty:
    raise ValueError("No rows left after number cleaning.")

# ============================================================
# KAUM CLEANING + FILTER
# ============================================================

df["kaum_clean"] = df["kategori_kaum"].apply(clean_kaum)

print("\nKAUM VALUES FOUND AFTER CLEANING:")
print(sorted(df["kaum_clean"].astype(str).unique()))

before_kaum = len(df)

if KEEP_KAUM != ["ALL"]:
    keep_set = set(normalize_filter_value(x) for x in KEEP_KAUM)
    df = df[df["kaum_clean"].isin(keep_set)].copy()

after_kaum = len(df)

print(f"After kaum filter         : {format_count(after_kaum)} kept ({format_count(before_kaum - after_kaum)} removed)")

if df.empty:
    raise ValueError("No rows left after kaum filtering.")

# ============================================================
# SIKAP CLEANING + FILTER
# ============================================================

if "sikap" in df.columns:
    df["sikap_clean"] = df["sikap"].apply(normalize_sikap)

    print("\nSIKAP VALUES FOUND AFTER CLEANING:")
    print(sorted(df["sikap_clean"].astype(str).unique()))

else:
    df["sikap_clean"] = ""

if SIKAP_FILTER:
    before_sikap = len(df)
    sikap_set = set(normalize_sikap(x) for x in SIKAP_FILTER)

    print(f"\nSIKAP FILTER USED: {sorted(sikap_set)}")

    df = df[df["sikap_clean"].isin(sikap_set)].copy()

    after_sikap = len(df)

    print(f"After sikap filter        : {format_count(after_sikap)} kept ({format_count(before_sikap - after_sikap)} removed)")

if df.empty:
    raise ValueError("No rows left after sikap filtering.")

# ============================================================
# PARTY FILTER
# ============================================================

if PARTY_FILTER:
    before_party = len(df)
    party_set = set(str(x).strip().upper() for x in PARTY_FILTER)

    df["party_clean"] = df["party"].astype(str).str.strip().str.upper()

    df = df[df["party_clean"].isin(party_set)].copy()

    after_party = len(df)

    print(f"After party filter        : {format_count(after_party)} kept ({format_count(before_party - after_party)} removed)")

if df.empty:
    raise ValueError("No rows left after party filtering.")

# ============================================================
# AGE FILTER
# ============================================================

if AGE_FILTER is not None:
    min_age, max_age = AGE_FILTER

    before_age = len(df)

    df["umur_num"] = pd.to_numeric(df["umur"], errors="coerce")

    if max_age is None:
        df = df[df["umur_num"] >= min_age].copy()
    else:
        df = df[
            (df["umur_num"] >= min_age) &
            (df["umur_num"] <= max_age)
        ].copy()

    after_age = len(df)

    print(f"After age filter          : {format_count(after_age)} kept ({format_count(before_age - after_age)} removed)")

if df.empty:
    raise ValueError("No rows left after age filtering.")

# ============================================================
# WA FORMAT COLUMNS
# ============================================================

df["code"] = df.apply(
    lambda x: get_code(x.get("kategori_kaum", ""), x.get("jantina", "")),
    axis=1
)

df["First Name"] = df.apply(format_first_name, axis=1)

before_first_name = len(df)
df = df[df["First Name"].astype(str).str.strip() != ""].copy()
after_first_name = len(df)

print(f"After First Name filter   : {format_count(after_first_name)} kept ({format_count(before_first_name - after_first_name)} removed)")

df["Last Name"] = df["code"]

if df.empty:
    raise ValueError("No rows left after WA formatting.")

# ============================================================
# BUILD GROUP LABELS
# ============================================================

df = build_labels(df)

if "AGE" in GROUP_LEVELS:
    before_age_group = len(df)

    df = df[df["__AGE_LABEL"].notna()].copy()

    after_age_group = len(df)

    print(f"After age range grouping  : {format_count(after_age_group)} kept ({format_count(before_age_group - after_age_group)} removed)")

if df.empty:
    raise ValueError("No rows left after grouping setup.")

# ============================================================
# WRITE OUTPUT
# ============================================================

summary_lines = [
    "SUMMARY",
    "",
    f"JUMLAH AKHIR = {format_count(len(df))}",
    "",
    "CONFIG",
    f"INPUT_LEVEL = {INPUT_LEVEL}",
    f"GROUP_LEVELS = {GROUP_LEVELS}",
    f"LAST_GROUP_AS_FOLDER = {LAST_GROUP_AS_FOLDER}",
    f"SPLIT_BY_KAUM = {SPLIT_BY_KAUM}",
    f"SPLIT_BY_GENDER_CODE = {SPLIT_BY_GENDER_CODE}",
    f"KEEP_KAUM = {KEEP_KAUM}",
    f"SIKAP_FILTER = {SIKAP_FILTER}",
    f"PARTY_FILTER = {PARTY_FILTER}",
    f"AGE_FILTER = {AGE_FILTER}",
    f"DEDUP_BY_PHONE = {DEDUP_BY_PHONE}",
    f"DEDUP_BY_NOKP = {DEDUP_BY_NOKP}",
    f"READ_ALL_SHEETS = {READ_ALL_SHEETS}",
    "",
    "KAUM NOTE",
    "FINAL KAUM BUCKETS = MELAYU, CINA, INDIA, LAIN-LAIN",
    "ORANG ASLI, BUMI, AND ANY OTHER KAUM = LAIN-LAIN",
    "",
    "SIKAP NOTE",
    "KELABU = any value containing KELABU, including KELABU-BARU and KELABU-LAMA",
    "",
    "OUTPUT COUNT",
    ""
]

label_cols, sort_cols = get_group_columns()

sort_existing = [c for c in sort_cols if c in df.columns]

if sort_existing:
    df = df.sort_values(sort_existing).copy()

split_col = None

if SPLIT_BY_GENDER_CODE:
    split_col = "code"

elif SPLIT_BY_KAUM:
    split_col = "kaum_clean"

total_files_created = 0

# ============================================================
# CASE 1 — NO GROUPING
# ============================================================

if not label_cols:
    base_folder = OUTPUT_ROOT
    os.makedirs(base_folder, exist_ok=True)

    if split_col:
        split_groups = df.groupby(split_col, dropna=False, sort=False)

        for split_value, split_group in tqdm(split_groups, desc="Writing files"):
            if split_group.empty and not CREATE_EMPTY_FILES:
                continue

            export_df = build_wa_df(split_group)

            out_file = f"{safe_file_label(split_value)}.xlsx"
            out_path = os.path.join(base_folder, out_file)

            save_xlsx_no_style(export_df, out_path)

            total_files_created += 1
            summary_lines.append(f"{safe_file_label(split_value)} = {format_count(len(split_group))}")

            if PRINT_SAVED_LINES:
                print(f"Saved: {out_path} ({format_count(len(split_group))} rows)")

    else:
        export_df = build_wa_df(df)

        out_file = "OUTPUT.xlsx"
        out_path = os.path.join(base_folder, out_file)

        save_xlsx_no_style(export_df, out_path)

        total_files_created += 1
        summary_lines.append(f"OUTPUT = {format_count(len(df))}")

        if PRINT_SAVED_LINES:
            print(f"Saved: {out_path} ({format_count(len(df))} rows)")

# ============================================================
# CASE 2 — GROUPING
# ============================================================

else:
    grouped = df.groupby(label_cols, dropna=False, sort=False)

    for group_key, group in tqdm(grouped, desc="Writing outputs"):
        if not isinstance(group_key, tuple):
            group_key = (group_key,)

        group_labels = [safe_file_label(x) for x in group_key]

        if len(group_labels) == 1:
            if LAST_GROUP_AS_FOLDER:
                folder_parts = group_labels
                file_base = group_labels[-1]
            else:
                folder_parts = []
                file_base = group_labels[-1]

        else:
            if LAST_GROUP_AS_FOLDER:
                folder_parts = group_labels
                file_base = group_labels[-1]
            else:
                folder_parts = group_labels[:-1]
                file_base = group_labels[-1]

        out_folder = os.path.join(OUTPUT_ROOT, *folder_parts)
        os.makedirs(out_folder, exist_ok=True)

        display_path = " / ".join(group_labels)
        summary_lines.append(f"{display_path} = {format_count(len(group))}")

        if split_col:
            split_groups = group.groupby(split_col, dropna=False, sort=False)

            for split_value, split_group in split_groups:
                if split_group.empty and not CREATE_EMPTY_FILES:
                    continue

                split_label = safe_file_label(split_value)
                export_df = build_wa_df(split_group)

                out_file = f"{file_base} {split_label}.xlsx"
                out_file = sanitize(out_file)
                out_path = os.path.join(out_folder, out_file)

                save_xlsx_no_style(export_df, out_path)

                total_files_created += 1
                summary_lines.append(f"  {split_label} = {format_count(len(split_group))}")

                if PRINT_SAVED_LINES:
                    print(f"Saved: {out_path} ({format_count(len(split_group))} rows)")

        else:
            if group.empty and not CREATE_EMPTY_FILES:
                continue

            export_df = build_wa_df(group)

            out_file = f"{file_base}.xlsx"
            out_file = sanitize(out_file)
            out_path = os.path.join(out_folder, out_file)

            save_xlsx_no_style(export_df, out_path)

            total_files_created += 1

            if PRINT_SAVED_LINES:
                print(f"Saved: {out_path} ({format_count(len(group))} rows)")

        summary_lines.append("")

# ============================================================
# SUMMARY FILE
# ============================================================

summary_text = "\n".join(summary_lines)

summary_path = os.path.join(OUTPUT_ROOT, "SUMMARY.txt")

with open(summary_path, "w", encoding="utf-8") as f:
    f.write(summary_text)

if PRINT_FULL_SUMMARY_ON_SCREEN:
    print("\n" + "=" * 60)
    print(summary_text)
    print("=" * 60)

else:
    print("\nDONE WRITING FILES")
    print("=" * 40)
    print(f"Final rows          : {format_count(len(df))}")
    print(f"Excel files created : {format_count(total_files_created)}")
    print(f"Summary saved       : {summary_path}")
    print("=" * 40)

# ============================================================
# ZIP + DOWNLOAD
# ============================================================

zip_path = shutil.make_archive(ZIP_NAME, "zip", OUTPUT_ROOT)
files.download(zip_path)

print("ZIP downloaded.")